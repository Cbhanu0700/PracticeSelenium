import openpyxl

file="D:\\Book1.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet3"]

rows=sheet.max_row
cells=sheet.max_column

for row in range(1,rows+1):
    for cell in range(1,cells+1):
        print(sheet.cell(row,cell).value,end="   ")
    print()