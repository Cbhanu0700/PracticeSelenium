import openpyxl



path="D:\\Book1.xlsx"

workbook=openpyxl.load_workbook(path)
sheet=workbook["Sheet2"]
# same data
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value="row"
#
# workbook.save(path)
#multiple data


sheet.cell(1,1).value=123
sheet.cell(1,2).value="Smith"
sheet.cell(2,1).value=456
sheet.cell(2,2).value="JOhn"
sheet.cell(3,1).value=678
sheet.cell(3,2).value="David"

workbook.save(path)



